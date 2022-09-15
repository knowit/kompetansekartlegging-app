import asyncio
import concurrent.futures
import threading
import json
from tempfile import NamedTemporaryFile
import boto3
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Color, Border, Font
from openpyxl.chart.radar_chart import RadarChart
from openpyxl.chart.bar_chart import BarChart
from openpyxl.chart.reference import Reference
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter, FORMULAE
from openpyxl.writer.excel import save_virtual_workbook
from dateutil import parser
from datetime import datetime
from os import environ


region = environ.get("AWS_REGION")

db_client = boto3.client("dynamodb")
cog_client = boto3.client("cognito-idp")
s3_client = boto3.client(
    's3', region_name=region, endpoint_url=f'https://s3.{region}.amazonaws.com',
)

env = environ.get("ENV")
sourceName = environ.get("SOURCE_NAME")
userPoolId = environ.get("USER_POOL_ID")
bucketName = environ.get("EXCEL_BUCKET")

formDefTable = f"FormDefinition-{sourceName}-{env}"
categoryTable = f"Category-{sourceName}-{env}"
questionTable = f"Question-{sourceName}-{env}"
userFormTable = f"UserForm-{sourceName}-{env}"
questionAnswerTable = f"QuestionAnswer-{sourceName}-{env}"

def sortByCreatedAt(obj):
    return parser.parse(obj["createdAt"])

def fetch_user_answers(username, formDef):
    userFormRes = db_client.query(
        TableName=userFormTable,
        IndexName="byCreatedAt",
        KeyConditionExpression="#owner = :username",
        FilterExpression="formDefinitionID=:formdef",
        ExpressionAttributeNames= {
            "#owner": "owner"
        },
        ExpressionAttributeValues={
            ":username": {"S": username},
            ":formdef": {"S": formDef}
        },
    )
    userForms = []
    for userForm in userFormRes["Items"]:
        attributes = {}
        for key in userForm.keys():
            attributes[key] = userForm[key]['S'] if 'S' in userForm[key].keys() else userForm[key]['N']
        userForms.append(attributes)

    userForms.sort(key=sortByCreatedAt, reverse=True)
    if len(userForms) <= 0: return {"username": username, "userFormId":"", "answers":{}}
    currentUserForm = userForms[0]

    answerRes = db_client.query(
        TableName=questionAnswerTable,
        IndexName="byUserForm",
        KeyConditionExpression="userFormID = :userFormId",
        ExpressionAttributeValues={
            ":userFormId": {"S": currentUserForm["id"]}
        },
    )
    answers = {}
    if "LastEvaluatedKey" in answerRes.keys():
        print("LastEvaluatedKey included in keys")
    for answer in answerRes["Items"]:
        attributes = {}
        for key in answer.keys():
            attributes[key] = answer[key][list(answer[key].keys())[0]]# if 'S' in answer[key].keys() elif '' in answer[key].keys() answer[key]['N']
        answers[attributes["questionID"]] = attributes
    return {"username": username, "userFormId": currentUserForm["id"], "answers": answers}


async def fetch_org_data(orgid):
    later = datetime.now()
    formDefsResponse = db_client.query(
            TableName=formDefTable,
            IndexName="byOrganizationByCreatedAt",
            KeyConditionExpression="organizationID = :org",
            ExpressionAttributeValues={":org": {"S": orgid}},
            Select="ALL_ATTRIBUTES"
        )
    formDefs = []
    for item in formDefsResponse["Items"]:
        attributes = {}
        for key in item.keys():
            attributes[key] = item[key]['S']
        formDefs.append(attributes)

    currentFormDef = formDefs[-1]["id"]
    print("Time passed fetching form defs:", datetime.now() - later)
    later = datetime.now()
    catResponse = db_client.query(
        TableName=categoryTable,
        IndexName="byFormDefinition",
        KeyConditionExpression="formDefinitionID = :formDef",
        ExpressionAttributeValues={":formDef": {"S": currentFormDef}},
    )

    categories = []
    for item in catResponse["Items"]:
        attributes = {}
        for key in item.keys():
            attributes[key] = item[key]['S'] if 'S' in item[key].keys() else item[key]['N']
        categories.append(attributes)
    print("Time passed fetching categories:", datetime.now() - later)
    later = datetime.now()

    questionResponse = db_client.query(
        TableName=questionTable,
        IndexName="byFormDefinition",
        Select="ALL_ATTRIBUTES",
        KeyConditionExpression="formDefinitionID = :formDef",
        ExpressionAttributeValues={":formDef": {"S": currentFormDef}},
    )

    questions = []
    for item in questionResponse["Items"]:
        attributes = {}
        for key in item.keys():
            attributes[key] = item[key]['S'] if 'S' in item[key].keys() else item[key]['N']
        # print(attributes)
        questions.append(attributes)
    print("Time passed fetching questions:", datetime.now() - later)
    later = datetime.now()

    users = []
    userRes = cog_client.list_users_in_group(
        UserPoolId=userPoolId,
        GroupName=orgid
    )
    for user in userRes["Users"]:
        attributes = {}
        for attribute in user["Attributes"]:
            attributes[attribute["Name"]] = attribute["Value"]
        user["Attributes"] = attributes
        users.append(user)

    next_token = None
    if ("NextToken" in userRes.keys()):
        next_token = userRes["NextToken"]
    while next_token:
        userRes = cog_client.list_users_in_group(
            UserPoolId=userPoolId,
            GroupName=orgid,
            NextToken=next_token
        )
        for user in userRes["Users"]:
            attributes = {}
            for attribute in user["Attributes"]:
                attributes[attribute["Name"]] = attribute["Value"]
            user["Attributes"] = attributes
            users.append(user)
        next_token = None
        if ("NextToken" in userRes.keys()):
            next_token = userRes["NextToken"]
    print("Time passed fetching users:", datetime.now() - later)
    later = datetime.now()
    
    mappedUsers = []
    answerTasks = []
    # with concurrent.futures.ThreadPoolExecutor(max_workers=3) as executor:

    for user in users:
        userAnswers = asyncio.to_thread(fetch_user_answers, user["Username"], currentFormDef)
        answerTasks.append(userAnswers)

    answers = await asyncio.gather(*answerTasks)

    for a in answers:
        userAnswers = a
        mappedUsers.append({"username": userAnswers["username"], "email": userAnswers["username"], "answers": userAnswers["answers"], "userFormId": userAnswers["userFormId"]})
    
    print("Time passed fetching answers:", datetime.now() - later)

    return formDefs, categories, questions, mappedUsers

def handler(event, context):
    print(event)
    later = datetime.now()
    groups = event["requestContext"]["authorizer"]["claims"]["cognito:groups"].split(",")
    isAdmin = False
    orgid = ""
    print("Hello", groups)
    for group in groups:
        print(group)
        roles = group.split("0")
        if len(roles) > 1 and roles[1] == "admin":
            isAdmin = True
            orgid = roles[0]

    if isAdmin:
        print("User is admin")
        book = asyncio.run(make_workbook(orgid))
        print("Time passed make_workbook:", datetime.now() - later)
        with NamedTemporaryFile(mode="w+b") as temp:
            book.save(temp.name)
            temp.seek(0)
            print("Time passed save workbook:", datetime.now() - later)
            # stream = temp.read()
            key = f"{orgid}_report_{datetime.now().isoformat(sep='-',timespec='minutes')}.xlsx"
            s3_client.put_object(Bucket=bucketName, Key=key, Body=temp)
            presigned_url = s3_client.generate_presigned_url('get_object',
                                                        Params={'Bucket': bucketName,
                                                                'Key': key},
                                                        ExpiresIn=5*60)
            print("Time passed total:", datetime.now() - later)
            return {
                "statusCode": 200,
                "headers": {
                    "Content-Type": "application/vnd.ms-excel",
                    'Access-Control-Allow-Headers': 'Content-Type',
                    "Access-Control-Allow-Origin": "*",
                    "Access-Control-Allow-Credentials": True
                },
                "body": presigned_url#f"https://{bucketName}.s3.amazonaws.com/{key}"
            }
    return {
                "statusCode": 200,
                "headers": {
                    "Content-Type": "application/vnd.ms-excel",
                    'Access-Control-Allow-Headers': 'Content-Type',
                    "Access-Control-Allow-Origin": "*",
                    "Access-Control-Allow-Credentials": True
                },
                "body": json.dumps("User is not Admin")}

def add_user_row(data_sheet, aggregateSheet, row, user, questions, questionPositions, questionBlockStart, categories, questionCategoryMap):

    cell = data_sheet.cell(row, 1, user["username"])
    data_sheet.merge_cells(start_row=row, end_row=row, start_column=1, end_column=3)
    for question in questions:
        if not question["id"] in user["answers"].keys():
            cell = data_sheet.cell(row, questionPositions[question["id"]][0], 0)
            cell = data_sheet.cell(row, questionPositions[question["id"]][1], 0)
            continue
        answer = user["answers"][question["id"]]
        if "knowledge" in answer.keys():
            if float(answer["knowledge"]) >= 0:
                cell = data_sheet.cell(row, questionPositions[question["id"]][0], float(answer["knowledge"]))
                cell.fill = PatternFill(patternType= "solid",fgColor=Color(rgb="AAAAFF", tint=(1 - float(answer["knowledge"])/5.0)), bgColor="aa000000")
            else:
                cell = data_sheet.cell(row, questionPositions[question["id"]][0], 0)
        if "motivation" in answer.keys():
            if float(answer["motivation"]) >= 0:
                cell = data_sheet.cell(row, questionPositions[question["id"]][1], float(answer["motivation"]))
                cell.fill = PatternFill(patternType= "solid",fgColor=Color(rgb="AAAAFF", tint=(1 - float(answer["motivation"])/5.0)))
                continue
            else:
                cell = data_sheet.cell(row, questionPositions[question["id"]][1], 0)
                continue
        if "type" in question.keys() and question["type"] == "customScaleLabels":
            if "customScaleValue" in answer.keys():
                cell = data_sheet.cell(row, questionPositions[question["id"]][0], float(answer["customScaleValue"]))
                cell = data_sheet.cell(row, questionPositions[question["id"]][1], float(answer["customScaleValue"]))
            else:
                cell = data_sheet.cell(row, questionPositions[question["id"]][0], 0)
                cell = data_sheet.cell(row, questionPositions[question["id"]][1], 0)
            continue
        print("Dominus")
    catCol = questionBlockStart
    aggCol = questionBlockStart
    aggregateSheet.cell(row + 7, aggCol - 1, user["username"])
    for category in categories:
        endCol = catCol + len(questionCategoryMap[category["id"]]) - 1
        aggregateSheet.cell(row + 7, aggCol, f"=AVERAGE({data_sheet.title}!${get_column_letter(catCol)}{row}:${get_column_letter(endCol)}{row})")
        aggregateSheet.cell(row + 7, aggCol + 1, f"=MEDIAN({data_sheet.title}!${get_column_letter(catCol)}{row}:${get_column_letter(endCol)}{row})")
        aggregateSheet.cell(row + 7, aggCol + 2, f"=PERCENTILE({data_sheet.title}!${get_column_letter(catCol)}{row}:${get_column_letter(endCol)}{row}, 0.95)")
        aggregateSheet.cell(row + 7, aggCol + 3, f"=MAX({data_sheet.title}!${get_column_letter(catCol)}{row}:${get_column_letter(endCol)}{row})")
        catCol += len(questionCategoryMap[category["id"]])
        aggCol += 4

async def make_workbook(orgid):
    later = datetime.now()
    formDefs, categories, questions, users = await fetch_org_data(orgid)
    print("Time passed fetching data:", datetime.now() - later)
    later = datetime.now()
    book = Workbook()
    questions.sort(key=lambda x: x["categoryID"])

    data_sheet = book.active
    data_sheet.title = "data"

    aggregateSheet = book.create_sheet("Aggergates")
    chartSheet = book.create_sheet("Charts")

    questionBlockStart = 4 #Excel is 1 indexed

    questionCategoryMap = {}
    for question in questions:
        if not question["categoryID"] in questionCategoryMap.keys():
            questionCategoryMap[question["categoryID"]] = []
        questionCategoryMap[question["categoryID"]].append(question)


    col = questionBlockStart
    grayColor = Color(rgb="000000", tint=0.6)

    cell = data_sheet.cell(1, questionBlockStart, "Kompetanse")
    cell.fill = PatternFill(patternType="solid", fgColor=grayColor)
    cell.alignment = Alignment(horizontal="center")
    data_sheet.merge_cells(start_row=1, end_row=1, start_column=questionBlockStart, end_column=questionBlockStart+len(questions)-1)

    cell = data_sheet.cell(1, questionBlockStart + len(questions), "Motivasjon")
    cell.fill = PatternFill(patternType="solid", fgColor=grayColor)
    cell.alignment = Alignment(horizontal="center")
    data_sheet.merge_cells(start_row=1, end_row=1, start_column=questionBlockStart + len(questions), end_column=questionBlockStart + len(questions) + len(questions) - 1)

    questionPositions = {}
    userStartRow = 5
    aggCol = questionBlockStart

    for category in categories:
        categoryColor = Color(rgb="AAAAFF", tint=0.2)
        questionColor = Color(rgb="AAAAFF", tint=0.5)
        cell = data_sheet.cell(2, col, category["text"])
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(patternType= "solid",fgColor=categoryColor)
        data_sheet.merge_cells(start_row=2, end_row=2, start_column=col, end_column=col + len(questionCategoryMap[category["id"]])-1)

        cell = data_sheet.cell(2, col + len(questions), category["text"])
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(patternType= "solid",fgColor=categoryColor)
        data_sheet.merge_cells(start_row=2, end_row=2, start_column=col+len(questions), end_column=col + len(questions) + len(questionCategoryMap[category["id"]])-1)

        cell = aggregateSheet.cell(userStartRow + 5, aggCol, category["text"])
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(patternType= "solid",fgColor=categoryColor)
        aggregateSheet.merge_cells(start_row=userStartRow + 5, end_row=userStartRow + 5, start_column=aggCol, end_column=aggCol + 3)

        cell = aggregateSheet.cell(userStartRow + 6, aggCol    , "Average")
        cell.fill = PatternFill(patternType= "solid",fgColor=grayColor)
        cell = aggregateSheet.cell(userStartRow + 6, aggCol + 1, "Median")
        cell.fill = PatternFill(patternType= "solid",fgColor=grayColor)
        cell = aggregateSheet.cell(userStartRow + 6, aggCol + 2, "Percentile (95)")
        cell.fill = PatternFill(patternType= "solid",fgColor=grayColor)
        cell = aggregateSheet.cell(userStartRow + 6, aggCol + 3, "Max")
        cell.fill = PatternFill(patternType= "solid",fgColor=grayColor)
        aggCol += 4


        for question in questionCategoryMap[category["id"]]:
            data_sheet.column_dimensions[get_column_letter(col)].width = 4
            data_sheet.column_dimensions[get_column_letter(col + len(questions))].width = 4

            cell = data_sheet.cell(3, col, question["topic"])
            cell.alignment = Alignment(text_rotation=90)
            cell.fill = PatternFill(patternType= "solid",fgColor=questionColor)
            cell.border = Border(outline=True)
            # col += 1
            cell = data_sheet.cell(3, col + len(questions), question["topic"])
            cell.alignment = Alignment(text_rotation=90)
            cell.fill = PatternFill(patternType= "solid",fgColor=questionColor)
            cell.border = Border(outline=True)

            cell = data_sheet.cell(4, col, question["id"])
            cell.fill = PatternFill(patternType= "solid",fgColor=grayColor)
            cell.border = Border(outline=True)
            # col += 1
            cell = data_sheet.cell(4, col + len(questions), question["id"])
            cell.fill = PatternFill(patternType= "solid",fgColor=grayColor)
            cell.border = Border(outline=True)

            questionPositions[question["id"]] = (col, col+len(questions))
            col += 1

    print("Time passed adding headers to data sheet:", datetime.now() - later)
    later = datetime.now()

    noAnswer = []
    row = userStartRow
    for user in users:
        if not user["answers"].keys(): 
            noAnswer.append(user["username"])
            continue
        add_user_row(data_sheet, aggregateSheet, int(row), user, questions, questionPositions, questionBlockStart, categories, questionCategoryMap)

        row += 1
    print("Time passed adding rows to data/aggregate sheet:", datetime.now() - later)
    later = datetime.now()

    aggCol = questionBlockStart
    # cell = aggregateSheet.cell(row + 7, aggCol - 1, "Total:")
    # cell.font = Font(bold=True)
    chartRow = 3
    chartSheet.cell(chartRow - 1, 2, "Average")
    chartSheet.cell(chartRow - 1, 3, "Median")
    chartSheet.cell(chartRow - 1, 4, "Max")

    for category in categories:
        #endCol = catCol + len(questionCategoryMap[category["id"]]) - 1
        chartSheet.cell(chartRow, 1, f"{category['text']})")
        chartSheet.cell(chartRow, 2, f"=AVERAGE({aggregateSheet.title}!${get_column_letter(aggCol)}{userStartRow + 7}:${get_column_letter(aggCol)}{row + 6})")
        chartSheet.cell(chartRow, 3, f"=MEDIAN({aggregateSheet.title}!${get_column_letter(aggCol+1)}{userStartRow + 7}:${get_column_letter(aggCol+1)}{row + 6})")
        chartSheet.cell(chartRow, 4, f"=MAX({aggregateSheet.title}!${get_column_letter(aggCol+3)}{userStartRow + 7}:${get_column_letter(aggCol+3)}{row + 6})")
        #catCol += len(questionCategoryMap[category["id"]])
        aggCol += 4
        chartRow += 1

    barGraph = BarChart()
    barGraph.type="bar"
    barGraph.title = "Category Summaries"
    barData = Reference(chartSheet, min_col=2, min_row=2, max_col=4, max_row=chartRow-1)
    barCategories = Reference(chartSheet, min_col=1, min_row=3, max_col=1, max_row=chartRow-1)
    barGraph.add_data(barData, titles_from_data=True)
    barGraph.set_categories(barCategories)
    chartSheet.add_chart(barGraph, f"{get_column_letter(7)}3")

    # for col, question in enumerate(questions):
    #     cell = chartSheet.cell(23, col + 2, question["topic"])
    # filtr = f"=FILTER({data_sheet.title}!{get_column_letter(1)}{userStartRow}:{get_column_letter(questionBlockStart + len(questions) - 1)}{row};{data_sheet.title}!{get_column_letter(1)}{userStartRow}:{get_column_letter(1)}{row}=A23)"
    # cell = chartSheet.cell(22, 1, f'For å bruke radar grafen, sett inn denne funksjonen i celle A24:')
    # cell = chartSheet.cell(22, 2, f'"{filtr}"')

    # radarChart = RadarChart(radarStyle="filled")
    # radar_data = Reference(chartSheet, min_col=2, max_col= 2+len(questions)-1, min_row=23, max_row=24)
    # radar_cols = Reference(chartSheet, min_col=2, max_col= 2+len(questions)-1, min_row=23, max_row=23)
    # radarChart.add_data(radar_data)
    # radarChart.set_categories(radar_cols)
    # chartSheet.add_chart(radarChart, f"{get_column_letter(15)}1")

    # userNames=[]
    # for user in users:
    #     userNames.append(user["username"])
    # userDataVal = DataValidation(type="list", formula1=f'={data_sheet.title}!A{userStartRow}:A{row}', allow_blank=True)
    # cell = chartSheet.cell(23, 1, userNames[0])
    
    # userDataVal.add(cell)
    # chartSheet.add_data_validation(userDataVal)


    print("Time passed adding charts to chart sheet:", datetime.now() - later)
    later = datetime.now()
    

    aggregateSheet.cell(3, 3, f"Average")
    aggregateSheet.cell(4, 3, f"Median")
    for question in questions:
        knowledgePos = questionPositions[question['id']][0]
        motivationPos = questionPositions[question['id']][1]
        aggregateSheet.cell(2, knowledgePos, f"{question['text']}")
        aggregateSheet.cell(2, motivationPos, f"{question['text']}")
        aggregateSheet.cell(3, knowledgePos, f"=AVERAGE({data_sheet.title}!${get_column_letter(knowledgePos)}{userStartRow}:${get_column_letter(knowledgePos)}{row})")
        aggregateSheet.cell(3, motivationPos, f"=AVERAGE({data_sheet.title}!${get_column_letter(motivationPos)}{userStartRow}:${get_column_letter(motivationPos)}{row})")
        aggregateSheet.cell(4, knowledgePos, f"=MEDIAN({data_sheet.title}!${get_column_letter(knowledgePos)}{userStartRow}:${get_column_letter(knowledgePos)}{row})")
        aggregateSheet.cell(4, motivationPos, f"=MEDIAN({data_sheet.title}!${get_column_letter(motivationPos)}{userStartRow}:${get_column_letter(motivationPos)}{row})")
    print("Time passed adding aggregates to aggregate sheet:", datetime.now() - later)

    return book

# Uncomment code below and fill in ENV, SourceName and UserpoolId to run locally

# env = ""
# sourceName = ""
# userPoolId = ""

# formDefTable = f"FormDefinition-{sourceName}-{env}"
# categoryTable = f"Category-{sourceName}-{env}"
# questionTable = f"Question-{sourceName}-{env}"
# userFormTable = f"UserForm-{sourceName}-{env}"
# questionAnswerTable = f"QuestionAnswer-{sourceName}-{env}"
# async def main():
#     later = datetime.now()
#     book = await make_workbook("knowitobjectnet")
#     book.save("testy.xlsx")
#     print(datetime.now() - later)

# asyncio.run(main())
"""
TODO: Write the rest of the code.
TODO: Figure out how to send an Excel file from Lambda :)
"""